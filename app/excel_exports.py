"""Professional Excel exports for operational, attendance, DTR, and project reports."""
from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timezone
from io import BytesIO
import re
from typing import Any, Iterable
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import DEFAULT_TIMEZONE
from app.models import (
    DailyAttendance,
    DTRDailyLine,
    Freelancer,
    LeaveRequest,
    MonthlyDTR,
    OvertimeClaim,
    PortalProject,
)
from app.performance_reporting import build_performance_dashboard, build_project_reports
from app.portal_project_service import project_overview_rows, task_overview_rows, team_assignment_rows
from app.task_time_reporting import build_task_time_utilization
from app.work_order_service import live_work_rows

NAVY = "17365D"
BLUE = "2F6FA5"
LIGHT_BLUE = "EAF3FB"
LIGHT_GRAY = "F3F6F9"
WHITE = "FFFFFF"
GREEN = "E7F6ED"
YELLOW = "FFF4CC"
RED = "FDECEC"
PURPLE = "EFE7FA"
TEXT = "243447"
BORDER_COLOR = "D6DEE6"
THIN = Side(style="thin", color=BORDER_COLOR)


EXCEL_ZH = {
    "Export Summary": "匯出摘要", "All Tasks": "所有任務", "Projects": "專案", "Team Availability": "團隊可用狀態",
    "Monthly Attendance": "月度出勤", "Monthly DTR Summary": "月度 DTR 摘要", "DTR Daily Details": "DTR 每日明細",
    "Leave Requests": "請假申請", "Overtime Claims": "加班申請", "Performance": "績效", "Project Reports": "專案報表",
    "Monthly Project Time": "每月專案工時", "Project Time Utilization": "專案工時利用率", "Task Time Details": "任務工時明細",
    "Project Work Time Health": "專案工時概況", "Project Time by Member": "專案成員工時", "Monthly Breakdown": "每月工時明細",
    "BIM PORTAL — FREELANCER EXPORT PACKAGE": "BIM PORTAL — 自由工作者匯出資料包",
    "Generated {value}": "產生時間 {value}", "Metric": "指標", "Value": "數值", "Purpose": "用途", "Scope": "範圍",
    "Selected Month": "選定月份", "Open Tasks": "未結案任務", "Completed Tasks": "已完成任務", "Unassigned Open Tasks": "未指派未結案任務",
    "Active Team Members": "啟用團隊成員", "DTR Records": "DTR 紀錄", "Portal records": "Portal 紀錄",
    "Monthly attendance and DTR period": "月度出勤與 DTR 期間", "Current project register": "目前專案清冊",
    "Complete task register": "完整任務清冊", "Tasks requiring action": "需要處理的任務", "Completed task history": "已完成任務歷程",
    "Tasks requiring assignment": "需要指派的任務", "Current team availability population": "目前團隊可用人員",
    "Generated DTR records for selected month": "選定月份已產生的 DTR 紀錄",
    "ALL TASKS": "所有任務", "Complete task register including current assignment, schedule, progress, and Quality Score.": "完整任務清冊，包含目前指派、時程、進度及品質分數。",
    "Task ID": "任務 ID", "Project": "專案", "Project Engineer": "專案工程師", "Task": "任務", "Description": "說明", "Assigned Member": "指派成員",
    "Status": "狀態", "Priority": "優先順序", "Discipline": "專業別", "Progress %": "進度 %", "Quality Score": "品質分數", "Start Date": "開始日期",
    "Deadline": "截止日期", "Completed Date": "完成日期", "Updated At": "更新時間",
    "PROJECT REGISTER": "專案清冊", "Current project-level status, progress, team coverage, and active workload.": "目前各專案狀態、進度、團隊配置與進行中工作量。",
    "Project ID": "專案 ID", "Project Category": "專案類別", "Members": "成員數", "Active Tasks": "進行中任務",
    "TEAM AVAILABILITY": "團隊可用狀態", "Green = available, blue = working now, yellow = assigned, red = overdue.": "綠色＝可用、藍色＝工作中、黃色＝已指派、紅色＝逾期。",
    "Member": "成員", "Code": "代碼", "Join Date": "加入日期", "Availability": "可用狀態", "Current Project": "目前專案", "Working Task": "目前任務",
    "Overdue": "逾期", "Completed": "已完成", "Assignment Status": "指派狀態", "Overdue": "逾期", "Working Now": "工作中", "Assigned": "已指派", "Available": "可用",
    "MONTHLY ATTENDANCE — {value}": "月度出勤 — {value}", "Daily attendance records for the selected month.": "選定月份的每日出勤紀錄。",
    "Date": "日期", "Active": "啟用", "Time In": "上班時間", "Time Out": "下班時間", "Break Minutes": "休息分鐘", "Rendered": "實際工時", "Late": "遲到",
    "Undertime": "不足工時", "Overtime": "加班", "Review Status": "審查狀態", "Locked": "已鎖定", "Yes": "是", "No": "否",
    "MONTHLY DTR SUMMARY — {value}": "月度 DTR 摘要 — {value}", "Generated DTR status and monthly totals for all members.": "所有成員的 DTR 狀態與月度合計。",
    "DTR ID": "DTR ID", "Calendar Days": "日曆天數", "Scheduled Workdays": "排定工作日", "Present": "出勤", "Late Days": "遲到天數", "Absent": "缺勤",
    "Leave": "請假", "Holiday": "假日", "Rest Days": "休息日", "Incomplete": "不完整", "Approved OT": "核准加班", "Task Entries": "任務筆數", "Task Time": "任務工時", "Task Review": "任務審查",
    "DTR DAILY DETAILS — {value}": "DTR 每日明細 — {value}", "Day-by-day attendance and task totals supporting the monthly DTR.": "支援月度 DTR 的逐日出勤與任務工時明細。",
    "Day": "星期", "Day Type": "日期類型", "Attendance Status": "出勤狀態", "Scheduled Start": "排定開始", "Scheduled End": "排定結束", "Potential OT": "可能加班", "Task Summary": "任務摘要", "Notes": "備註",
    "LEAVE REQUESTS — {value}": "請假申請 — {value}", "Leave requests submitted for the selected month.": "選定月份提交的請假申請。",
    "Type": "類型", "Requested": "申請時數", "Approved": "核准時數", "Reason": "原因", "Review Reason": "審查原因", "Submitted At": "提交時間",
    "OVERTIME CLAIMS — {value}": "加班申請 — {value}", "Overtime claims and approved time for the selected month.": "選定月份的加班申請與核准時數。",
    "Potential": "可能加班", "Comp Credit": "補休額度", "Work Description": "工作說明", "Missing Time Out Reason": "缺少下班時間原因",
    "PERFORMANCE OVERVIEW": "績效總覽", "Task output, delivery, and Quality Score as calculated.": "任務產出、交付與品質分數。",
    "Rank": "排名", "Total Tasks": "任務總數", "Completion %": "完成率 %", "Rated Tasks": "已評分任務", "Measured Tasks": "可衡量任務", "On-time %": "準時率 %", "Average Days": "平均天數",
    "PROJECT REPORTS — {value}": "專案報表 — {value}", "Project delivery, active work, overdue tasks, Quality Score, and logged time.": "專案交付、進行中工作、逾期任務、品質分數與登錄工時。",
    "Delivered": "已交付", "Rated": "已評分", "Logged Hours": "登錄工時",
    "MONTHLY PROJECT WORK TIME — {value}": "每月專案工時 — {value}", "Logged production time by month, project, and member. Project Total repeats for each contributing member for easy filtering.": "依月份、專案與成員列示登錄的生產工時；專案總工時會在各貢獻成員列重複顯示，方便篩選。",
    "Month": "月份", "Project Code": "專案代碼", "Project Total": "專案總工時", "Member Code": "成員代碼", "Member Time": "成員工時",
    "PROJECT TIME UTILIZATION": "專案工時利用率", "Planned time is Start Date through Deadline. Utilization time includes production time plus saved review time; completed tasks without production hours use Start Date through Completion Date as the production estimate.": "計畫工時以開始日期至截止日期計算；利用工時包含生產工時與已儲存的審查工時。已完成但無生產工時的任務，使用開始日期至完成日期估算生產工時。",
    "Tasks": "任務數", "Tasks with Utilization": "有利用率資料的任務", "Planned Time": "計畫工時", "Utilization Time": "利用工時", "Production Recorded Time": "已登錄生產工時", "Review Time": "審查工時", "Recorded Total": "登錄總工時", "Completion Estimate Time": "完成估算工時", "All-time Project Hours": "專案歷史總工時", "Recorded Without Plan": "無計畫工時的登錄時間", "Remaining / Overrun": "剩餘／超支", "Time Budget Used %": "工時預算使用率 %", "Unlinked Time": "未連結工時",
    "TASK TIME UTILIZATION DETAILS": "任務工時利用率明細", "Utilization includes recorded production time plus saved review time. Completed tasks without production hours use Start Date through Completion Date as the production estimate.": "利用工時包含登錄生產工時與已儲存的審查工時。已完成但無生產工時的任務，以開始日期至完成日期估算生產工時。",
    "Time Source": "工時來源", "Included": "納入計算", "Calculation": "計算方式", "Contributors": "貢獻成員",
    "PROJECT WORK TIME HEALTH — {value}": "專案工時概況 — {value}",
    "Work-time health shows how much production time was logged to each project in the selected reporting period. It does not grade productivity; it shows recorded activity and contribution coverage.": "工時概況顯示所選期間內各專案登錄的生產工時。此欄位不評分生產力，只呈現實際登錄活動與成員貢獻範圍。",
    "Reporting Period": "報告期間", "Total Work Time": "總工作工時", "Members Logging Time": "有登錄工時成員", "Top Contributor": "主要貢獻成員", "Top Contributor Time": "主要貢獻工時", "Share of All Project Time %": "占全部專案工時 %", "Active Logging Months": "有工時月份數", "Last Logged Date": "最後登錄日期", "Work Time Health": "工時狀態", "Time Logged": "已有工時", "No Logged Time": "無登錄工時",
    "PROJECT TIME BY MEMBER — {value}": "專案成員工時 — {value}", "Selected-period project totals broken down by every member who logged production time.": "將所選期間的專案總工時依每位登錄生產工時的成員拆分。", "Share of Project %": "占專案工時 %",
    "MONTHLY PROJECT TIME BREAKDOWN — {value}": "每月專案工時明細 — {value}", "Calendar-month breakdown of logged production time for the selected reporting period.": "所選報告期間內依日曆月份拆分的生產工時。",
    "Monthly": "月度", "12 Months": "12 個月", "All Time": "全部期間",
    "Not calculated — complete Start and Deadline": "未計算 — 請完成開始日期與截止日期", "Completion-date estimate + review": "完成日期估算 + 審查", "Completion-date estimate": "完成日期估算", "Unlinked recorded time": "未連結的登錄工時", "Production + review": "生產 + 審查", "Review time": "審查工時", "Recorded production time": "已登錄生產工時",
    "COMPLETED": "已完成", "IN_PROGRESS": "進行中", "FOR_REVIEW": "待審查", "PENDING": "待處理", "PENDING_PLAN": "待計畫核准", "PENDING_FINAL": "待最終核准", "CANCELLED": "已取消", "FINALIZED": "已定稿", "APPROVED": "已核准", "REJECTED": "已拒絕", "NOT_STARTED": "未開始", "ON_HOLD": "暫停", "ACTIVE": "啟用", "INACTIVE": "停用", "Unassigned": "未指派",
}

EXCEL_DATA_TRANSLATABLE = {
    "COMPLETED", "IN_PROGRESS", "FOR_REVIEW", "PENDING", "PENDING_PLAN", "PENDING_FINAL",
    "CANCELLED", "FINALIZED", "APPROVED", "REJECTED", "NOT_STARTED", "ON_HOLD", "ACTIVE", "INACTIVE",
    "Available", "Working Now", "Assigned", "Overdue", "Yes", "No", "Unassigned",
    "Time Logged", "No Logged Time", "Not calculated — complete Start and Deadline",
    "Completion-date estimate + review", "Completion-date estimate", "Unlinked recorded time",
    "Production + review", "Review time", "Recorded production time",
}


def _xl(locale: str, text: str, **values: object) -> str:
    localized = EXCEL_ZH.get(text, text) if str(locale or "en") == "zh_TW" else text
    try:
        return localized.format(**values)
    except (KeyError, ValueError):
        return localized

def _xl_status(locale: str, value: object) -> str:
    raw = str(value or "")
    key = raw.strip().upper()
    return _xl(locale, key) if key in EXCEL_ZH else raw

def _xl_sheet(wb: Workbook, locale: str, name: str):
    return wb.create_sheet(_xl(locale, name)[:31])


_ZH_MONTHS = {
    "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, "June": 6,
    "July": 7, "August": 8, "September": 9, "October": 10, "November": 11, "December": 12,
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}

def _zh_month_text(text: str) -> str:
    month_pattern = "|".join(sorted((re.escape(key) for key in _ZH_MONTHS), key=len, reverse=True))
    def repl(match):
        return f"{match.group(2)}年{_ZH_MONTHS[match.group(1)]}月"
    return re.sub(rf"\b({month_pattern})\s+(\d{{4}})\b", repl, text)

def _xl_dynamic(locale: str, text: str) -> str:
    if str(locale or "en") != "zh_TW":
        return text
    exact = EXCEL_ZH.get(text)
    if exact is not None:
        return _zh_month_text(exact)
    for pattern, translated in EXCEL_ZH.items():
        if "{value}" not in pattern:
            continue
        prefix, suffix = pattern.split("{value}", 1)
        if text.startswith(prefix) and text.endswith(suffix):
            middle_end = len(text) - len(suffix) if suffix else len(text)
            value = text[len(prefix):middle_end]
            return _zh_month_text(translated.format(value=value))
    return _zh_month_text(text)


def _localize_workbook(wb: Workbook, locale: str) -> None:
    if str(locale or "en") != "zh_TW":
        return
    for sheet in wb.worksheets:
        original_title = sheet.title
        sheet.title = _xl(locale, original_title)[:31]
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                # Rows 1, 2 and 4 are generated title/subtitle/header rows.
                # Data rows are left untouched except for known categorical
                # values so project/member/user-entered text is never translated.
                if cell.row in {1, 2, 4}:
                    cell.value = _xl_dynamic(locale, cell.value)
                elif original_title == "Export Summary" and cell.column in {1, 3, 4}:
                    cell.value = _xl_dynamic(locale, cell.value)
                elif cell.value in EXCEL_DATA_TRANSLATABLE:
                    cell.value = _xl_dynamic(locale, cell.value)


def _project_period_label(period: str, month_key: str, fallback: str, locale: str) -> str:
    if str(locale or "en") != "zh_TW":
        return fallback
    if period == "all":
        return "全部期間"
    try:
        selected = datetime.strptime(month_key, "%Y-%m").date().replace(day=1)
    except ValueError:
        return fallback
    if period == "month":
        return f"{selected.year}年{selected.month}月"
    if period == "12m":
        index = selected.year * 12 + (selected.month - 1) - 11
        start = date(index // 12, index % 12 + 1, 1)
        return f"{start.year}年{start.month}月 – {selected.year}年{selected.month}月"
    return fallback


def _zone(name: str | None) -> ZoneInfo:
    try:
        return ZoneInfo(str(name or DEFAULT_TIMEZONE))
    except ZoneInfoNotFoundError:
        return ZoneInfo(DEFAULT_TIMEZONE)


def _local_time(value: datetime | None, timezone_name: str | None) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(_zone(timezone_name)).strftime("%H:%M")


def _month_bounds(month_key: str) -> tuple[str, date, date]:
    try:
        start = datetime.strptime(str(month_key or ""), "%Y-%m").date().replace(day=1)
    except ValueError:
        start = date.today().replace(day=1)
    days = monthrange(start.year, start.month)[1]
    end = date(start.year + (1 if start.month == 12 else 0), 1 if start.month == 12 else start.month + 1, 1)
    return start.strftime("%Y-%m"), start, end


def _duration(minutes: int | None) -> str:
    value = max(0, int(minutes or 0))
    hours, remainder = divmod(value, 60)
    return f"{hours}h {remainder:02d}m"


def _new_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def _title(sheet, title: str, subtitle: str, last_column: int) -> int:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(size=16, bold=True, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sub = sheet.cell(2, 1, subtitle)
    sub.font = Font(size=9, italic=True, color=TEXT)
    sub.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    sub.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[2].height = 22
    return 4


def _header(sheet, row: int, headers: list[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row, column, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    sheet.row_dimensions[row].height = 28


def _write_rows(
    sheet,
    *,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: Iterable[Iterable[Any]],
    widths: list[int] | None = None,
    status_column: int | None = None,
    due_column: int | None = None,
) -> None:
    header_row = _title(sheet, title, subtitle, len(headers))
    _header(sheet, header_row, headers)
    today = date.today()
    for row_index, values in enumerate(rows, header_row + 1):
        values = list(values)
        for col_index, value in enumerate(values, 1):
            cell = sheet.cell(row_index, col_index, value)
            cell.font = Font(size=9, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        fill = None
        if status_column is not None and status_column <= len(values):
            status = str(values[status_column - 1] or "").strip().upper()
            if status == "COMPLETED":
                fill = GREEN
            elif status in {"IN_PROGRESS", "FOR_REVIEW", "PENDING", "PENDING_PLAN", "PENDING_FINAL"}:
                fill = YELLOW
        if due_column is not None and due_column <= len(values):
            due_value = values[due_column - 1]
            if isinstance(due_value, str) and due_value and due_value != "—":
                try:
                    due = date.fromisoformat(due_value)
                except ValueError:
                    due = None
                status = str(values[status_column - 1] or "").strip().upper() if status_column else ""
                if due and due < today and status not in {"COMPLETED", "CANCELLED", "FINALIZED"}:
                    fill = RED
        if fill:
            for cell in sheet[row_index]:
                cell.fill = PatternFill("solid", fgColor=fill)
    sheet.freeze_panes = f"A{header_row + 1}"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{sheet.max_row}"
    if widths:
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
    else:
        for index in range(1, len(headers) + 1):
            longest = max(
                (len(str(sheet.cell(row, index).value or "")) for row in range(1, min(sheet.max_row, 200) + 1)),
                default=10,
            )
            sheet.column_dimensions[get_column_letter(index)].width = max(10, min(35, longest + 2))


def _add_summary_sheet(wb: Workbook, month_key: str, database: Session) -> None:
    sheet = wb.create_sheet("Export Summary")
    tasks = task_overview_rows(database, status_mode="all", limit=1000)
    projects = project_overview_rows(database, limit=500)
    team = team_assignment_rows(database)
    dtrs = list(database.scalars(select(MonthlyDTR).where(MonthlyDTR.month_key == month_key)).all())
    cards = [
        ("Selected Month", month_key),
        ("Projects", len(projects)),
        ("All Tasks", len(tasks)),
        ("Open Tasks", sum(str(row["status"]).upper() not in {"COMPLETED", "CANCELLED"} for row in tasks)),
        ("Completed Tasks", sum(str(row["status"]).upper() == "COMPLETED" for row in tasks)),
        ("Unassigned Open Tasks", sum(str(row.get("assignees") or "") == "Unassigned" and str(row["status"]).upper() not in {"COMPLETED", "CANCELLED"} for row in tasks)),
        ("Active Team Members", len(team)),
        ("DTR Records", len(dtrs)),
    ]
    _title(sheet, "BIM PORTAL — FREELANCER EXPORT PACKAGE", f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", 4)
    _header(sheet, 4, ["Metric", "Value", "Purpose", "Scope"])
    purpose = {
        "Selected Month": "Monthly attendance and DTR period",
        "Projects": "Current project register",
        "All Tasks": "Complete task register",
        "Open Tasks": "Tasks requiring action",
        "Completed Tasks": "Completed task history",
        "Unassigned Open Tasks": "Tasks requiring assignment",
        "Active Team Members": "Current team availability population",
        "DTR Records": "Generated DTR records for selected month",
    }
    for idx, (label, value) in enumerate(cards, 5):
        sheet.cell(idx, 1, label)
        sheet.cell(idx, 2, value)
        sheet.cell(idx, 3, purpose[label])
        sheet.cell(idx, 4, "Portal records")
        for cell in sheet[idx]:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(idx, 1).font = Font(bold=True, color=NAVY)
        sheet.cell(idx, 2).font = Font(bold=True, size=13, color=BLUE)
    for col, width in enumerate([26, 16, 42, 24], 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A5"


def _add_tasks_sheet(wb: Workbook, database: Session) -> None:
    rows = task_overview_rows(database, status_mode="all", limit=1000)
    _write_rows(
        wb.create_sheet("All Tasks"),
        title="ALL TASKS",
        subtitle="Complete task register including current assignment, schedule, progress, and Quality Score.",
        headers=[
            "Task ID", "Project", "Project Engineer", "Task", "Description", "Assigned Member",
            "Status", "Priority", "Discipline", "Progress %", "Quality Score", "Start Date",
            "Deadline", "Completed Date", "Updated At",
        ],
        rows=[[
            row["id"], row["project_name"], row["project_engineer"], row["title"], row["description"],
            row["assignees"], row["status"], row["priority"], row["discipline"], row["progress"],
            row["quality_score"] if row["quality_score"] is not None else "", row["start_date"],
            row["due_date"], row["completion_date"],
            row["updated_at"].strftime("%Y-%m-%d %H:%M") if row.get("updated_at") else "",
        ] for row in rows],
        widths=[10, 28, 22, 30, 42, 28, 18, 12, 16, 12, 14, 13, 13, 15, 18],
        status_column=7,
        due_column=13,
    )


def _add_projects_sheet(wb: Workbook, database: Session) -> None:
    rows = project_overview_rows(database, limit=500)
    _write_rows(
        wb.create_sheet("Projects"),
        title="PROJECT REGISTER",
        subtitle="Current project-level status, progress, team coverage, and active workload.",
        headers=["Project ID", "Project", "Project Category", "Project Engineer", "Status", "Priority", "Progress %", "Members", "Active Tasks", "Deadline"],
        rows=[[
            row["id"], row["name"], row["project_category"], row["project_engineer"], row["status"], row["priority"],
            row["progress"], row["member_count"], row["active_task_count"], row["deadline"],
        ] for row in rows],
        widths=[11, 32, 18, 24, 16, 12, 13, 11, 13, 14],
        status_column=5,
        due_column=10,
    )


def _add_team_sheet(wb: Workbook, database: Session) -> None:
    live = {int(row["freelancer_id"]): row for row in live_work_rows(database)}
    rows = []
    for row in team_assignment_rows(database):
        current = live.get(int(row["freelancer_id"]))
        if int(row["overdue_task_count"] or 0) > 0:
            state = "Overdue"
        elif current:
            state = "Working Now"
        elif int(row["active_task_count"] or 0) > 0:
            state = "Assigned"
        else:
            state = "Available"
        rows.append([
            row["name"], row["code"], row.get("join_date", "—"), state,
            current["project_name"] if current else "", current["task_title"] if current else "",
            row["project_count"], row["active_task_count"], row["overdue_task_count"], row["completed_task_count"],
            row["assignment_status"],
        ])
    _write_rows(
        wb.create_sheet("Team Availability"),
        title="TEAM AVAILABILITY",
        subtitle="Green = available, blue = working now, yellow = assigned, red = overdue.",
        headers=["Member", "Code", "Join Date", "Availability", "Current Project", "Working Task", "Projects", "Active Tasks", "Overdue", "Completed", "Assignment Status"],
        rows=rows,
        widths=[26, 14, 13, 16, 28, 32, 10, 12, 10, 11, 28],
    )
    sheet = wb["Team Availability"]
    for row in range(5, sheet.max_row + 1):
        state = str(sheet.cell(row, 4).value or "")
        fill = {"Available": GREEN, "Working Now": LIGHT_BLUE, "Assigned": YELLOW, "Overdue": RED}.get(state)
        if fill:
            for cell in sheet[row]:
                cell.fill = PatternFill("solid", fgColor=fill)


def _add_attendance_sheet(wb: Workbook, database: Session, month_key: str) -> None:
    _, start, end = _month_bounds(month_key)
    results = list(database.execute(
        select(DailyAttendance, Freelancer)
        .join(Freelancer, Freelancer.id == DailyAttendance.freelancer_id)
        .where(DailyAttendance.attendance_date >= start, DailyAttendance.attendance_date < end)
        .order_by(DailyAttendance.attendance_date, Freelancer.full_name)
    ).all())
    _write_rows(
        wb.create_sheet("Monthly Attendance"),
        title=f"MONTHLY ATTENDANCE — {month_key}",
        subtitle="Daily attendance records for the selected month.",
        headers=["Date", "Member", "Code", "Join Date", "Active", "Time In", "Time Out", "Break Minutes", "Rendered", "Late", "Undertime", "Overtime", "Status", "Review Status", "Locked"],
        rows=[[
            record.attendance_date.isoformat(), freelancer.full_name, freelancer.freelancer_code,
            freelancer.join_date.isoformat() if freelancer.join_date else "—", "Yes" if freelancer.is_active else "No",
            _local_time(record.time_in_utc, freelancer.timezone_name), _local_time(record.time_out_utc, freelancer.timezone_name),
            record.break_minutes, _duration(record.rendered_minutes), _duration(record.late_minutes),
            _duration(record.undertime_minutes), _duration(record.overtime_minutes), record.status,
            record.review_status, "Yes" if record.is_locked else "No",
        ] for record, freelancer in results],
        widths=[13, 26, 14, 13, 10, 11, 11, 14, 13, 12, 13, 12, 16, 16, 10],
        status_column=13,
    )


def _add_dtr_sheets(wb: Workbook, database: Session, month_key: str) -> None:
    records = list(database.execute(
        select(MonthlyDTR, Freelancer)
        .join(Freelancer, Freelancer.id == MonthlyDTR.freelancer_id)
        .where(MonthlyDTR.month_key == month_key)
        .order_by(Freelancer.full_name)
    ).all())
    _write_rows(
        wb.create_sheet("Monthly DTR Summary"),
        title=f"MONTHLY DTR SUMMARY — {month_key}",
        subtitle="Generated DTR status and monthly totals for all members.",
        headers=["DTR ID", "Member", "Code", "Join Date", "Status", "Calendar Days", "Scheduled Workdays", "Present", "Late Days", "Absent", "Leave", "Holiday", "Rest Days", "Incomplete", "Rendered", "Late", "Undertime", "Approved OT", "Task Entries", "Task Time", "Task Review"],
        rows=[[
            dtr.id, freelancer.full_name, freelancer.freelancer_code,
            freelancer.join_date.isoformat() if freelancer.join_date else "—", dtr.status,
            dtr.calendar_days, dtr.scheduled_workdays, dtr.present_days, dtr.late_days, dtr.absent_days,
            dtr.leave_days, dtr.holiday_days, dtr.rest_days, dtr.incomplete_days, _duration(dtr.rendered_minutes),
            _duration(dtr.late_minutes), _duration(dtr.undertime_minutes), _duration(dtr.approved_overtime_minutes),
            dtr.daily_task_entries, _duration(dtr.daily_task_minutes), dtr.task_review_status,
        ] for dtr, freelancer in records],
        widths=[10, 26, 14, 13, 13, 13, 17, 10, 11, 10, 10, 10, 11, 12, 13, 12, 13, 13, 12, 13, 15],
        status_column=5,
    )
    dtr_ids = [int(dtr.id) for dtr, _ in records]
    names = {int(dtr.id): (freelancer.full_name, freelancer.freelancer_code, freelancer.timezone_name) for dtr, freelancer in records}
    lines = list(database.scalars(
        select(DTRDailyLine)
        .where(DTRDailyLine.monthly_dtr_id.in_(tuple(dtr_ids)) if dtr_ids else DTRDailyLine.id.is_(None))
        .order_by(DTRDailyLine.attendance_date, DTRDailyLine.monthly_dtr_id)
    ).all())
    _write_rows(
        wb.create_sheet("DTR Daily Details"),
        title=f"DTR DAILY DETAILS — {month_key}",
        subtitle="Day-by-day attendance and task totals supporting the monthly DTR.",
        headers=["Date", "Member", "Code", "Day", "Day Type", "Attendance Status", "Scheduled Start", "Scheduled End", "Time In", "Time Out", "Rendered", "Late", "Undertime", "Potential OT", "Approved OT", "Task Time", "Task Entries", "Task Summary", "Review Status", "Notes"],
        rows=[[
            line.attendance_date.isoformat(), names.get(int(line.monthly_dtr_id), ("—", "—", DEFAULT_TIMEZONE))[0],
            names.get(int(line.monthly_dtr_id), ("—", "—", DEFAULT_TIMEZONE))[1], line.day_name, line.day_type,
            line.attendance_status, line.scheduled_start_text, line.scheduled_end_text,
            _local_time(line.time_in_utc, names.get(int(line.monthly_dtr_id), ("", "", DEFAULT_TIMEZONE))[2]),
            _local_time(line.time_out_utc, names.get(int(line.monthly_dtr_id), ("", "", DEFAULT_TIMEZONE))[2]),
            _duration(line.rendered_minutes), _duration(line.late_minutes), _duration(line.undertime_minutes),
            _duration(line.potential_overtime_minutes), _duration(line.approved_overtime_minutes),
            _duration(line.task_minutes), line.task_entry_count, line.task_summary or "", line.attendance_review_status, line.notes or "",
        ] for line in lines],
        widths=[13, 26, 14, 11, 15, 18, 14, 14, 11, 11, 13, 12, 13, 13, 13, 13, 12, 42, 16, 36],
        status_column=6,
    )


def _add_requests_sheets(wb: Workbook, database: Session, month_key: str) -> None:
    _, start, end = _month_bounds(month_key)
    names = {int(row.id): row for row in database.scalars(select(Freelancer)).all()}
    leave = list(database.scalars(
        select(LeaveRequest)
        .where(LeaveRequest.leave_date >= start, LeaveRequest.leave_date < end)
        .order_by(LeaveRequest.leave_date, LeaveRequest.id)
    ).all())
    _write_rows(
        wb.create_sheet("Leave Requests"),
        title=f"LEAVE REQUESTS — {month_key}",
        subtitle="Leave requests submitted for the selected month.",
        headers=["Date", "Member", "Code", "Type", "Requested", "Approved", "Status", "Reason", "Review Reason", "Submitted At"],
        rows=[[
            item.leave_date.isoformat(), names.get(int(item.freelancer_id)).full_name if names.get(int(item.freelancer_id)) else "—",
            names.get(int(item.freelancer_id)).freelancer_code if names.get(int(item.freelancer_id)) else "—",
            item.leave_type, _duration(item.requested_minutes), _duration(item.approved_minutes), item.status,
            item.reason, item.review_reason or "", item.submitted_at.strftime("%Y-%m-%d %H:%M"),
        ] for item in leave],
        widths=[13, 26, 14, 18, 13, 13, 15, 40, 36, 18],
        status_column=7,
    )
    overtime = list(database.scalars(
        select(OvertimeClaim)
        .where(OvertimeClaim.attendance_date >= start, OvertimeClaim.attendance_date < end)
        .order_by(OvertimeClaim.attendance_date, OvertimeClaim.id)
    ).all())
    _write_rows(
        wb.create_sheet("Overtime Claims"),
        title=f"OVERTIME CLAIMS — {month_key}",
        subtitle="Overtime claims and approved time for the selected month.",
        headers=["Date", "Member", "Code", "Potential", "Requested", "Approved", "Comp Credit", "Status", "Work Description", "Missing Time Out Reason", "Review Reason", "Submitted At"],
        rows=[[
            item.attendance_date.isoformat(), names.get(int(item.freelancer_id)).full_name if names.get(int(item.freelancer_id)) else "—",
            names.get(int(item.freelancer_id)).freelancer_code if names.get(int(item.freelancer_id)) else "—",
            _duration(item.potential_minutes_snapshot), _duration(item.requested_minutes), _duration(item.approved_minutes),
            _duration(item.comp_leave_minutes_earned), item.status, item.work_description, item.missing_time_out_reason or "",
            item.review_reason or "", item.submitted_at.strftime("%Y-%m-%d %H:%M"),
        ] for item in overtime],
        widths=[13, 26, 14, 13, 13, 13, 14, 18, 42, 36, 36, 18],
        status_column=8,
    )


def _add_performance_sheets(wb: Workbook, database: Session, month_key: str) -> None:
    performance = build_performance_dashboard(database)
    _write_rows(
        wb.create_sheet("Performance"),
        title="PERFORMANCE OVERVIEW",
        subtitle="Task output, delivery, and Quality Score as calculated.",
        headers=["Rank", "Member", "Code", "Total Tasks", "Completed", "Active", "Completion %", "Rated Tasks", "Quality Score", "Measured Tasks", "On-time %", "Average Days"],
        rows=[[
            row.get("rank"), row.get("name"), row.get("code"), row.get("total_tasks"), row.get("completed_tasks"),
            row.get("active_tasks"), row.get("completion_rate"), row.get("rated_tasks"), row.get("average_quality"),
            row.get("measured_tasks"), row.get("delivery_rate"), row.get("average_days_label"),
        ] for row in performance.get("task_ranked", [])],
        widths=[9, 26, 14, 12, 12, 10, 13, 12, 14, 14, 12, 18],
    )
    reports = build_project_reports(database, period="month", month_key=month_key)
    _write_rows(
        wb.create_sheet("Project Reports"),
        title=f"PROJECT REPORTS — {reports.get('period_label', month_key)}",
        subtitle="Project delivery, active work, overdue tasks, Quality Score, and logged time.",
        headers=["Project", "Code", "Project Engineer", "Discipline", "Status", "Progress %", "Delivered", "Active", "Overdue", "Rated", "Quality Score", "Measured", "On-time %", "Logged Hours"],
        rows=[[
            row.get("name"), row.get("code"), row.get("project_engineer"), row.get("discipline"), row.get("status"),
            row.get("progress"), row.get("delivered_tasks"), row.get("active_tasks"), row.get("overdue_tasks"),
            row.get("rated_tasks"), row.get("average_quality"), row.get("measured_tasks"), row.get("on_time_rate"), row.get("logged_hours"),
        ] for row in reports.get("project_rows", [])],
        widths=[30, 16, 24, 16, 14, 12, 11, 10, 10, 10, 14, 11, 12, 14],
        status_column=5,
    )
    monthly_time_rows = []
    for project_row in reports.get("monthly_project_time_rows", []):
        members = project_row.get("members", []) or []
        if not members:
            monthly_time_rows.append([
                project_row.get("month"), project_row.get("project_name"), project_row.get("project_code"),
                _duration(project_row.get("total_minutes")), "—", "", "0h 00m",
            ])
            continue
        for member in members:
            monthly_time_rows.append([
                project_row.get("month"), project_row.get("project_name"), project_row.get("project_code"),
                _duration(project_row.get("total_minutes")), member.get("name"), member.get("code"),
                _duration(member.get("minutes")),
            ])
    _write_rows(
        wb.create_sheet("Monthly Project Time"),
        title=f"MONTHLY PROJECT WORK TIME — {reports.get('period_label', month_key)}",
        subtitle="Logged production time by month, project, and member. Project Total repeats for each contributing member for easy filtering.",
        headers=["Month", "Project", "Project Code", "Project Total", "Member", "Member Code", "Member Time"],
        rows=monthly_time_rows,
        widths=[12, 32, 18, 16, 28, 16, 16],
    )


def _add_utilization_sheets(wb: Workbook, database: Session) -> None:
    report = build_task_time_utilization(database)
    _write_rows(
        wb.create_sheet("Project Time Utilization"),
        title="PROJECT TIME UTILIZATION",
        subtitle="Planned time is Start Date through Deadline. Utilization time includes production time plus saved review time; completed tasks without production hours use Start Date through Completion Date as the production estimate.",
        headers=["Rank", "Project", "Code", "Status", "Tasks", "Tasks with Utilization", "Planned Time", "Utilization Time", "Production Recorded Time", "Review Time", "Recorded Total", "Completion Estimate Time", "All-time Project Hours", "Recorded Without Plan", "Remaining / Overrun", "Time Budget Used %", "Unlinked Time"],
        rows=[[
            row.get("actual_rank"), row.get("name"), row.get("code"), row.get("status"), row.get("task_count"),
            row.get("measured_task_count"), _duration(row.get("target_minutes")) if row.get("target_minutes") is not None else "—",
            _duration(row.get("measured_actual_minutes")), _duration(row.get("production_recorded_minutes")), _duration(row.get("review_minutes")),
            _duration(row.get("recorded_minutes")), _duration(row.get("completion_fallback_minutes")), _duration(row.get("actual_minutes")),
            _duration(row.get("excluded_actual_minutes")), row.get("variance_label"), row.get("utilization") if row.get("utilization") is not None else "",
            _duration(row.get("unlinked_minutes")),
        ] for row in report.get("projects", [])],
        widths=[9, 30, 16, 14, 10, 18, 14, 16, 20, 14, 16, 18, 18, 20, 20, 18, 14],
        status_column=4,
    )
    task_rows = []
    for project in report.get("projects", []):
        for row in project.get("rows", []):
            calculation = (
                f"{_duration(row.get('utilization_minutes'))} / {_duration(row.get('target_minutes'))} x 100"
                if row.get("included_in_utilization")
                else "Not calculated — complete Start and Deadline"
            )
            source = (
                "Completion-date estimate + review"
                if row.get("uses_completion_fallback") and row.get("review_minutes")
                else "Completion-date estimate"
                if row.get("uses_completion_fallback")
                else "Unlinked recorded time"
                if row.get("is_unlinked")
                else "Production + review"
                if row.get("production_recorded_minutes") and row.get("review_minutes")
                else "Review time"
                if row.get("review_minutes")
                else "Recorded production time"
            )
            task_rows.append([
                project.get("name"), row.get("title"), row.get("assignees"), row.get("status"), row.get("start_date"),
                row.get("deadline"), _duration(row.get("target_minutes")) if row.get("target_minutes") is not None else "—",
                _duration(row.get("production_recorded_minutes")), _duration(row.get("review_minutes")), _duration(row.get("utilization_minutes")), source,
                row.get("variance_label"), row.get("utilization") if row.get("utilization") is not None else "",
                "Yes" if row.get("included_in_utilization") else "No", calculation, row.get("contributors"),
            ])
    _write_rows(
        wb.create_sheet("Task Time Details"),
        title="TASK TIME UTILIZATION DETAILS",
        subtitle="Utilization includes recorded production time plus saved review time. Completed tasks without production hours use Start Date through Completion Date as the production estimate.",
        headers=["Project", "Task", "Assigned Member", "Status", "Start Date", "Deadline", "Planned Time", "Production Time", "Review Time", "Utilization Time", "Time Source", "Remaining / Overrun", "Time Budget Used %", "Included", "Calculation", "Contributors"],
        rows=task_rows,
        widths=[28, 34, 26, 15, 13, 13, 14, 16, 14, 16, 22, 20, 18, 11, 30, 42],
        status_column=4,
        due_column=6,
    )


def build_export_workbook(
    database: Session,
    *,
    month_key: str,
    locale: str = "en",
    include_tasks: bool = False,
    include_attendance: bool = False,
    include_dtr: bool = False,
    include_reports: bool = False,
    include_all: bool = False,
) -> bytes:
    month_key, _, _ = _month_bounds(month_key)
    wb = _new_workbook()
    _add_summary_sheet(wb, month_key, database)
    if include_all or include_tasks:
        _add_tasks_sheet(wb, database)
    if include_all or include_reports:
        _add_projects_sheet(wb, database)
        _add_team_sheet(wb, database)
        _add_performance_sheets(wb, database, month_key)
        _add_utilization_sheets(wb, database)
        _add_requests_sheets(wb, database, month_key)
    if include_all or include_attendance:
        _add_attendance_sheet(wb, database, month_key)
    if include_all or include_dtr:
        _add_dtr_sheets(wb, database, month_key)
    _localize_workbook(wb, locale)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def build_project_work_time_workbook(
    database: Session,
    *,
    period: str = "month",
    month_key: str = "",
    locale: str = "en",
) -> bytes:
    """Export project work-time health and member contribution for the selected report period."""
    report = build_project_reports(database, period=period, month_key=month_key)
    wb = _new_workbook()
    period_label = _project_period_label(
        str(report.get("period") or period),
        str(report.get("selected_month") or month_key),
        str(report.get("period_label") or month_key or ""),
        locale,
    )
    time_rows = {int(row["project_id"]): row for row in report.get("project_time_by_member_rows", [])}
    projects = project_overview_rows(database, limit=500)
    total_logged = sum(int(row.get("total_minutes") or 0) for row in time_rows.values())

    health_ranked_rows = []
    for project in projects:
        project_id = int(project["id"])
        activity = time_rows.get(project_id)
        total_minutes = int(activity.get("total_minutes") or 0) if activity else 0
        members = list(activity.get("members", []) or []) if activity else []
        top = members[0] if members else None
        health_ranked_rows.append((total_minutes, str(project.get("name") or ""), [
            project.get("name"),
            project.get("project_code") or project.get("code") or "",
            project.get("status") or "",
            project.get("progress") or 0,
            period_label,
            _duration(total_minutes),
            len(members),
            top.get("name") if top else "—",
            _duration(top.get("minutes")) if top else "0h 00m",
            round((total_minutes / total_logged * 100), 1) if total_logged else 0.0,
            int(activity.get("active_months") or 0) if activity else 0,
            activity.get("last_activity") or "—" if activity else "—",
            "Time Logged" if total_minutes > 0 else "No Logged Time",
        ]))
    health_ranked_rows.sort(key=lambda item: (-item[0], item[1].casefold()))
    health_rows = [item[2] for item in health_ranked_rows]

    _write_rows(
        wb.create_sheet("Project Work Time Health"),
        title=f"PROJECT WORK TIME HEALTH — {period_label}",
        subtitle="Work-time health shows how much production time was logged to each project in the selected reporting period. It does not grade productivity; it shows recorded activity and contribution coverage.",
        headers=["Project", "Project Code", "Status", "Progress %", "Reporting Period", "Total Work Time", "Members Logging Time", "Top Contributor", "Top Contributor Time", "Share of All Project Time %", "Active Logging Months", "Last Logged Date", "Work Time Health"],
        rows=health_rows,
        widths=[30, 17, 15, 12, 24, 17, 20, 26, 20, 24, 20, 16, 18],
        status_column=3,
    )

    member_rows = []
    for project_row in report.get("project_time_by_member_rows", []):
        members = project_row.get("members", []) or []
        if not members:
            member_rows.append([project_row.get("project_name"), project_row.get("project_code"), period_label, _duration(project_row.get("total_minutes")), "—", "", "0h 00m", 0.0])
            continue
        for member in members:
            member_rows.append([
                project_row.get("project_name"), project_row.get("project_code"), period_label,
                _duration(project_row.get("total_minutes")), member.get("name"), member.get("code"),
                _duration(member.get("minutes")), member.get("share_percent") or 0.0,
            ])
    _write_rows(
        wb.create_sheet("Project Time by Member"),
        title=f"PROJECT TIME BY MEMBER — {period_label}",
        subtitle="Selected-period project totals broken down by every member who logged production time.",
        headers=["Project", "Project Code", "Reporting Period", "Project Total", "Member", "Member Code", "Member Time", "Share of Project %"],
        rows=member_rows,
        widths=[30, 17, 24, 17, 26, 16, 17, 18],
    )

    monthly_rows = []
    for project_row in report.get("monthly_project_time_rows", []):
        members = project_row.get("members", []) or []
        if not members:
            monthly_rows.append([project_row.get("month"), project_row.get("project_name"), project_row.get("project_code"), _duration(project_row.get("total_minutes")), "—", "", "0h 00m"])
            continue
        for member in members:
            monthly_rows.append([project_row.get("month"), project_row.get("project_name"), project_row.get("project_code"), _duration(project_row.get("total_minutes")), member.get("name"), member.get("code"), _duration(member.get("minutes"))])
    _write_rows(
        wb.create_sheet("Monthly Breakdown"),
        title=f"MONTHLY PROJECT TIME BREAKDOWN — {period_label}",
        subtitle="Calendar-month breakdown of logged production time for the selected reporting period.",
        headers=["Month", "Project", "Project Code", "Project Total", "Member", "Member Code", "Member Time"],
        rows=monthly_rows,
        widths=[12, 30, 17, 17, 26, 16, 17],
    )

    _localize_workbook(wb, locale)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

