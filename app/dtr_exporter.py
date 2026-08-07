from datetime import timezone
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from app.database import SessionLocal
from app.hr_workflow import COMP_LEAVE_DAY_MINUTES
from app.models import (
    DTRCompLine,
    DTRDailyLine,
    DTRLeaveLine,
    DTRTaskLine,
    Freelancer,
    MonthlyDTR,
)
from app.payroll_engine import calculate_payroll_multiplier

NAVY = "17365D"
BLUE = "2F6FA5"
LIGHT_BLUE = "EAF3FB"
LIGHT_GRAY = "F3F6F9"
GREEN = "E7F6ED"
RED = "FDECEC"
GOLD = "FFF4CC"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="D9E1E8")


def duration_text(minutes: int, *, signed: bool = False) -> str:
    """Format time as hours and minutes without decimal-hour ambiguity."""
    total = int(minutes or 0)
    prefix = ""
    if signed:
        prefix = "+" if total > 0 else "-" if total < 0 else ""
    elif total < 0:
        prefix = "-"
    total = abs(total)
    hours, remainder = divmod(total, 60)
    return f"{prefix}{hours}h {remainder:02d}m"


def local_time_text(value, timezone_name: str, zone_getter) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(zone_getter(timezone_name)).strftime("%H:%M")


def title_row(sheet, title: str, last_column: int) -> None:
    sheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    cell = sheet.cell(1, 1, title)
    cell.font = Font(size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False


def header_row(sheet, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row, col, header)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=THIN_GRAY)


def autosize(sheet, widths: list[int]) -> None:
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width


def build_dtr_workbook(
    *,
    dtr: MonthlyDTR,
    freelancer: Freelancer,
    lines: Iterable[DTRDailyLine],
    generated_by: str,
    reviewed_by: Optional[str],
    finalized_by: Optional[str],
    zone_getter,
) -> bytes:
    attendance_lines = list(lines)
    with SessionLocal() as database:
        task_lines = list(
            database.scalars(
                select(DTRTaskLine)
                .where(DTRTaskLine.monthly_dtr_id == dtr.id)
                .order_by(DTRTaskLine.task_date, DTRTaskLine.id)
            ).all()
        )
        comp_lines = list(
            database.scalars(
                select(DTRCompLine)
                .where(DTRCompLine.monthly_dtr_id == dtr.id)
                .order_by(DTRCompLine.transaction_date, DTRCompLine.id)
            ).all()
        )
        leave_lines = list(
            database.scalars(
                select(DTRLeaveLine)
                .where(DTRLeaveLine.monthly_dtr_id == dtr.id)
                .order_by(DTRLeaveLine.leave_date, DTRLeaveLine.id)
            ).all()
        )

    comp_leave_taken_minutes = sum(
        int(line.comp_leave_minutes_used or 0) for line in leave_lines
    )
    worked_statuses = {
        "PRESENT",
        "LATE",
        "HOLIDAY_WORK",
        "REST_DAY_WORK",
        "WORKED_ON_LEAVE",
        "PARTIAL_LEAVE_WORK",
    }
    worked_days = len(
        {
            line.attendance_date
            for line in attendance_lines
            if line.attendance_status in worked_statuses
        }
    )
    approved_leave_days = len({line.leave_date for line in leave_lines})
    approved_leave_minutes = sum(max(0, int(line.duration_minutes or 0)) for line in leave_lines)

    payroll = calculate_payroll_multiplier(
        calendar_days=int(dtr.calendar_days or 0),
        approved_leave_minutes=approved_leave_minutes,
        comp_credit_minutes_available=max(comp_leave_taken_minutes, int(dtr.comp_leave_used_minutes or 0)),
        standard_day_minutes=COMP_LEAVE_DAY_MINUTES,
        absent_days=int(dtr.absent_days or 0),
    )
    payable_workday_equivalents = worked_days + payroll.comp_credit_days_applied

    wb = Workbook()
    summary = wb.active
    summary.title = "Finance Summary"
    title_row(summary, "BIM PORTAL MONTHLY DTR — FINANCE SUMMARY", 12)

    details = [
        ("Freelancer", freelancer.full_name, "Code", freelancer.freelancer_code),
        ("DTR Month", dtr.month_key, "Status", dtr.status),
        (
            "Schedule",
            dtr.schedule_name,
            "Work Hours",
            f"{dtr.scheduled_start_text}-{dtr.scheduled_end_text}",
        ),
        ("Timezone", dtr.timezone_name, "Generated By", generated_by),
        ("Reviewed By", reviewed_by or "—", "Finalized By", finalized_by or "—"),
        (
            "Task Review",
            dtr.task_review_status,
            "Finance Use",
            "Hourly leave-credit review — confidential salary is not stored",
        ),
    ]
    row = 3
    for left_label, left_value, right_label, right_value in details:
        summary[f"A{row}"] = left_label
        summary[f"B{row}"] = left_value
        summary[f"G{row}"] = right_label
        summary[f"H{row}"] = right_value
        summary.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        summary.merge_cells(start_row=row, start_column=8, end_row=row, end_column=12)
        for cell in (summary[f"A{row}"], summary[f"G{row}"]):
            cell.font = Font(bold=True, color=NAVY)
            cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        row += 1

    values = [
        ("TOTAL CALENDAR DAYS", payroll.calendar_days),
        ("DAYS PHYSICALLY WORKED", worked_days),
        ("APPROVED LEAVE TAKEN", f"{approved_leave_days} day(s) / {duration_text(payroll.approved_leave_minutes)}"),
        ("OVERTIME CREDIT APPLIED", duration_text(payroll.comp_credit_minutes_applied)),
        ("UNPAID LEAVE", duration_text(payroll.effective_unpaid_leave_minutes)),
        ("ABSENCE DEDUCTION", duration_text(payroll.absent_minutes)),
        ("PAYABLE WORKDAY EQUIVALENTS", payable_workday_equivalents),
        ("SALARY-COVERED CALENDAR DAYS", payroll.salary_coverage_display),
        ("EFFECTIVE PAYROLL DEDUCTION", payroll.deduction_display),
        ("PAYROLL TREATMENT", payroll.payroll_treatment_display),
        ("REST DAYS / WEEKENDS", int(dtr.rest_days or 0)),
        ("HOLIDAYS", int(dtr.holiday_days or 0)),
        ("UNRECORDED ABSENCE", int(dtr.absent_days or 0)),
        ("APPROVED OVERTIME", duration_text(dtr.approved_overtime_minutes)),
        ("OPENING COMP CREDITS", duration_text(dtr.comp_leave_opening_balance_minutes)),
        ("CREDITS EARNED FROM OT", duration_text(dtr.comp_leave_earned_minutes)),
        ("CREDITS APPLIED TO LEAVE", duration_text(dtr.comp_leave_used_minutes)),
        ("REMAINING COMP CREDITS", duration_text(dtr.comp_leave_closing_balance_minutes)),
        ("PENDING OT CLAIMS", int(dtr.pending_overtime_claims or 0)),
        ("PENDING LEAVE REQUESTS", int(dtr.pending_leave_requests or 0)),
        (
            "REVIEW STATUS",
            "READY"
            if dtr.incomplete_days == 0
            and dtr.pending_overtime_claims == 0
            and dtr.pending_leave_requests == 0
            else "REVIEW",
        ),
    ]
    start = 11
    highlighted = {
        "TOTAL CALENDAR DAYS",
        "DAYS PHYSICALLY WORKED",
        "OVERTIME CREDIT APPLIED",
        "PAYABLE WORKDAY EQUIVALENTS",
        "SALARY-COVERED CALENDAR DAYS",
    }
    for index, (label, value) in enumerate(values):
        column = 1 + (index % 4) * 3
        current_row = start + (index // 4) * 2
        summary.cell(current_row, column, label)
        summary.cell(current_row + 1, column, value)
        summary.merge_cells(
            start_row=current_row,
            start_column=column,
            end_row=current_row,
            end_column=column + 1,
        )
        summary.merge_cells(
            start_row=current_row + 1,
            start_column=column,
            end_row=current_row + 1,
            end_column=column + 1,
        )
        summary.cell(current_row, column).font = Font(bold=True, color=NAVY, size=9)
        summary.cell(current_row, column).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        summary.cell(current_row, column).alignment = Alignment(horizontal="center", wrap_text=True)
        summary.cell(current_row + 1, column).font = Font(bold=True, size=12)
        summary.cell(current_row + 1, column).alignment = Alignment(horizontal="center", wrap_text=True)
        if label in highlighted:
            summary.cell(current_row, column).fill = PatternFill("solid", fgColor=GREEN)
            summary.cell(current_row + 1, column).fill = PatternFill("solid", fgColor=GREEN)
            summary.cell(current_row + 1, column).font = Font(
                bold=True,
                size=16,
                color="176B3A",
            )
        elif label in {"UNPAID LEAVE", "ABSENCE DEDUCTION", "EFFECTIVE PAYROLL DEDUCTION"}:
            summary.cell(current_row, column).fill = PatternFill("solid", fgColor=GOLD)
            summary.cell(current_row + 1, column).fill = PatternFill("solid", fgColor=GOLD)

    autosize(summary, [21, 14, 4, 21, 14, 4, 21, 14, 4, 21, 14, 4])
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1

    note_row = start + ((len(values) + 3) // 4) * 2 + 1
    summary.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row + 2,
        end_column=12,
    )
    summary.cell(
        note_row,
        1,
        (
            f"FINANCE REVIEW: The freelancer physically worked {worked_days} day(s). "
            f"Approved leave taken: {approved_leave_days} day(s) / {duration_text(payroll.approved_leave_minutes)}. "
            f"Overtime credit applied: {duration_text(payroll.comp_credit_minutes_applied)}. "
            f"Unpaid leave: {duration_text(payroll.effective_unpaid_leave_minutes)}. "
            f"Absence deduction: {duration_text(payroll.absent_minutes)}. "
            f"Payable workday equivalents: {payable_workday_equivalents:.3f}. "
            f"Salary coverage: {payroll.salary_coverage_display}. "
            f"Payroll treatment: {payroll.payroll_treatment_display}. "
            "Approved overtime credit offsets approved leave minute-for-minute and never increases salary above the full monthly rate."
        ),
    )
    summary.cell(note_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    summary.cell(note_row, 1).fill = PatternFill("solid", fgColor=GOLD)
    summary.cell(note_row, 1).font = Font(bold=True, color=NAVY)

    attendance = wb.create_sheet("Daily Attendance")
    title_row(attendance, "MONTHLY DAILY TIME RECORD", 16)
    attendance_headers = [
        "Date",
        "Day",
        "Day Type",
        "Status",
        "Time In",
        "Time Out",
        "Rendered Time",
        "Late Min",
        "Undertime Min",
        "Potential OT",
        "Approved OT",
        "Comp Earned",
        "Comp Used",
        "Task Time",
        "Task Summary",
        "Notes",
    ]
    header_row(attendance, 3, attendance_headers)
    for row_number, line in enumerate(attendance_lines, 4):
        values = [
            line.attendance_date,
            line.day_name,
            line.day_type.replace("_", " ").title(),
            line.attendance_status.replace("_", " ").title(),
            local_time_text(line.time_in_utc, dtr.timezone_name, zone_getter),
            local_time_text(line.time_out_utc, dtr.timezone_name, zone_getter),
            duration_text(line.rendered_minutes),
            int(line.late_minutes or 0),
            int(line.undertime_minutes or 0),
            duration_text(line.potential_overtime_minutes),
            duration_text(line.approved_overtime_minutes),
            duration_text(line.comp_leave_earned_minutes),
            duration_text(line.comp_leave_used_minutes),
            duration_text(line.task_minutes),
            line.task_summary or "",
            line.notes or "",
        ]
        for column, value in enumerate(values, 1):
            attendance.cell(row_number, column, value).border = Border(bottom=THIN_GRAY)
            attendance.cell(row_number, column).alignment = Alignment(
                vertical="top",
                wrap_text=column in {15, 16},
            )
        attendance.cell(row_number, 1).number_format = "yyyy-mm-dd"
    autosize(attendance, [12, 11, 15, 20, 10, 10, 15, 10, 14, 15, 15, 15, 15, 13, 42, 36])
    attendance.freeze_panes = "A4"
    attendance.auto_filter.ref = f"A3:P{3 + len(attendance_lines)}"
    attendance.page_setup.orientation = "landscape"
    attendance.page_setup.fitToWidth = 1

    tasks = wb.create_sheet("Daily Tasks")
    title_row(tasks, "MONTHLY DAILY TASK REPORT", 10)
    task_headers = [
        "Date",
        "Project",
        "Discipline",
        "Task Description",
        "Accomplishment / Output",
        "Status",
        "Completion %",
        "Time Spent",
        "Notes",
        "Source ID",
    ]
    header_row(tasks, 3, task_headers)
    for row_number, line in enumerate(task_lines, 4):
        values = [
            line.task_date,
            line.project_name or "Project",
            line.discipline or "",
            line.task_description,
            line.accomplishment or "",
            line.task_status.replace("_", " ").title(),
            line.completion_percentage,
            duration_text(line.minutes_spent),
            line.notes or "",
            line.source_task_id,
        ]
        for column, value in enumerate(values, 1):
            tasks.cell(row_number, column, value).border = Border(bottom=THIN_GRAY)
            tasks.cell(row_number, column).alignment = Alignment(
                vertical="top",
                wrap_text=column in {4, 5, 9},
            )
        tasks.cell(row_number, 1).number_format = "yyyy-mm-dd"
        tasks.cell(row_number, 7).number_format = "0"
    autosize(tasks, [12, 28, 16, 42, 42, 15, 13, 14, 30, 11])
    tasks.freeze_panes = "A4"
    tasks.auto_filter.ref = f"A3:J{3 + len(task_lines)}"
    tasks.page_setup.orientation = "landscape"
    tasks.page_setup.fitToWidth = 1

    overtime = wb.create_sheet("Overtime Details")
    title_row(overtime, "OVERTIME ACCOUNTING DETAILS", 8)
    header_row(
        overtime,
        3,
        [
            "Date",
            "Attendance Status",
            "Potential OT",
            "Approved OT",
            "Approval Difference",
            "Comp Credit Earned",
            "Task Time",
            "Notes",
        ],
    )
    overtime_rows = [
        line
        for line in attendance_lines
        if int(line.potential_overtime_minutes or 0)
        or int(line.approved_overtime_minutes or 0)
        or int(line.comp_leave_earned_minutes or 0)
    ]
    for row_number, line in enumerate(overtime_rows, 4):
        values = [
            line.attendance_date,
            line.attendance_status.replace("_", " ").title(),
            duration_text(line.potential_overtime_minutes),
            duration_text(line.approved_overtime_minutes),
            duration_text(
                int(line.potential_overtime_minutes or 0)
                - int(line.approved_overtime_minutes or 0),
                signed=True,
            ),
            duration_text(line.comp_leave_earned_minutes),
            duration_text(line.task_minutes),
            line.notes or "",
        ]
        for column, value in enumerate(values, 1):
            overtime.cell(row_number, column, value).border = Border(bottom=THIN_GRAY)
            overtime.cell(row_number, column).alignment = Alignment(
                vertical="top",
                wrap_text=column == 8,
            )
        overtime.cell(row_number, 1).number_format = "yyyy-mm-dd"
    if not overtime_rows:
        overtime.merge_cells("A4:H5")
        overtime["A4"] = "No approved or potential overtime was recorded for this DTR month."
        overtime["A4"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    autosize(overtime, [12, 20, 18, 18, 22, 22, 14, 45])
    overtime.freeze_panes = "A4"

    comp = wb.create_sheet("OT and Comp Leave")
    title_row(comp, "OVERTIME AND COMPENSATORY LEAVE LEDGER", 6)
    header_row(
        comp,
        3,
        [
            "Date",
            "Transaction",
            "Credit Change",
            "Description",
            "Source ID",
            "Running Balance",
        ],
    )
    running = int(dtr.comp_leave_opening_balance_minutes or 0)
    comp.cell(4, 1, "Opening")
    comp.cell(4, 2, "Opening Balance")
    comp.cell(4, 6, duration_text(running))
    start_comp_row = 5
    if not comp_lines:
        comp.merge_cells("A5:F6")
        comp["A5"] = "No compensatory-credit ledger transactions were recorded for this DTR month."
        comp["A5"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row_number, line in enumerate(comp_lines, start_comp_row):
        running += int(line.amount_minutes or 0)
        values = [
            line.transaction_date,
            line.transaction_type.replace("_", " ").title(),
            duration_text(line.amount_minutes, signed=True),
            line.description,
            line.source_transaction_id,
            duration_text(running),
        ]
        for column, value in enumerate(values, 1):
            comp.cell(row_number, column, value).border = Border(bottom=THIN_GRAY)
            comp.cell(row_number, column).alignment = Alignment(
                vertical="top",
                wrap_text=column == 4,
            )
        comp.cell(row_number, 1).number_format = "yyyy-mm-dd"
    closing_row = start_comp_row + (len(comp_lines) if comp_lines else 2)
    comp.cell(closing_row, 1, "Closing")
    comp.cell(closing_row, 2, "Closing Balance")
    comp.cell(
        closing_row,
        6,
        duration_text(dtr.comp_leave_closing_balance_minutes),
    )
    autosize(comp, [12, 22, 20, 60, 12, 20])
    comp.freeze_panes = "A4"
    comp.auto_filter.ref = f"A3:F{closing_row}"

    leaves = wb.create_sheet("Leave Records")
    title_row(leaves, "APPROVED LEAVE RECORDS", 8)
    header_row(
        leaves,
        3,
        [
            "Date",
            "Leave Type",
            "Duration",
            "Comp Leave Applied",
            "Paid Classification",
            "Notes",
            "Source ID",
            "DTR Month",
        ],
    )
    if not leave_lines:
        leaves.merge_cells("A4:H5")
        leaves["A4"] = "No approved leave records were recorded for this DTR month."
        leaves["A4"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
    for row_number, line in enumerate(leave_lines, 4):
        values = [
            line.leave_date,
            line.leave_type.replace("_", " ").title(),
            duration_text(line.duration_minutes),
            duration_text(line.comp_leave_minutes_used),
            "Yes" if line.is_paid else "No",
            line.notes or "",
            line.source_leave_id,
            dtr.month_key,
        ]
        for column, value in enumerate(values, 1):
            leaves.cell(row_number, column, value).border = Border(bottom=THIN_GRAY)
            leaves.cell(row_number, column).alignment = Alignment(
                vertical="top",
                wrap_text=column == 6,
            )
        leaves.cell(row_number, 1).number_format = "yyyy-mm-dd"
    autosize(leaves, [12, 24, 14, 20, 18, 45, 12, 12])

    readiness = wb.create_sheet("Payroll Readiness")
    title_row(readiness, "FINANCE PAYROLL READINESS CHECKLIST", 4)
    expected_close = (
        int(dtr.comp_leave_opening_balance_minutes or 0)
        + int(dtr.comp_leave_earned_minutes or 0)
        - int(dtr.comp_leave_used_minutes or 0)
    )
    checks = [
        (
            "Attendance Complete",
            dtr.incomplete_days == 0 and dtr.scheduled_future_days == 0,
            "No incomplete or future attendance",
        ),
        (
            "Daily Tasks Complete",
            dtr.task_missing_days == 0
            and dtr.task_review_status in {"REVIEWED", "NOT_REQUIRED"},
            dtr.task_review_status,
        ),
        (
            "Overtime Approved",
            dtr.pending_overtime_claims == 0,
            f"{dtr.pending_overtime_claims} pending",
        ),
        (
            "Leave Approved",
            dtr.pending_leave_requests == 0,
            f"{dtr.pending_leave_requests} pending",
        ),
        (
            "Comp Ledger Balanced",
            expected_close == int(dtr.comp_leave_closing_balance_minutes or 0),
            (
                f"Expected {duration_text(expected_close)} / "
                f"Actual {duration_text(dtr.comp_leave_closing_balance_minutes)}"
            ),
        ),
    ]
    header_row(readiness, 3, ["Check", "Status", "Details", "Finance Decision"])
    for row_number, (label, ok, detail) in enumerate(checks, 4):
        readiness.cell(row_number, 1, label)
        readiness.cell(row_number, 2, "PASS" if ok else "FAIL")
        readiness.cell(row_number, 3, detail)
        readiness.cell(row_number, 2).fill = PatternFill(
            "solid",
            fgColor=GREEN if ok else RED,
        )
        for column in range(1, 5):
            readiness.cell(row_number, column).border = Border(bottom=THIN_GRAY)
    payroll_ready = all(ok for _, ok, _ in checks)
    readiness.cell(10, 1, "Payroll Ready")
    readiness.cell(10, 2, "YES" if payroll_ready else "NO")
    readiness.cell(10, 2).fill = PatternFill(
        "solid",
        fgColor=GREEN if payroll_ready else RED,
    )
    readiness.cell(
        10,
        4,
        "Finance may proceed to external payroll calculation."
        if payroll_ready
        else "Resolve failed checks before payroll processing.",
    )
    autosize(readiness, [28, 14, 42, 52])
    readiness.page_setup.fitToWidth = 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
