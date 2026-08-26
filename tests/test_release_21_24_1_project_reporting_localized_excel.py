from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.excel_exports import build_export_workbook, build_project_work_time_workbook
from app.models import DailyTask, Freelancer, PortalProject
from app.performance_reporting import build_project_reports

ROOT = Path(__file__).resolve().parents[1]


def _db():
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    return engine, Session(engine)


def _seed_project_time(db: Session):
    a = Freelancer(freelancer_code="F-001", full_name="Member One", timezone_name="Asia/Manila")
    b = Freelancer(freelancer_code="F-002", full_name="Member Two", timezone_name="Asia/Manila")
    project = PortalProject(project_code="P-100", name="Project Alpha", status="ACTIVE", priority="NORMAL", progress=25)
    db.add_all([a, b, project])
    db.flush()
    db.add_all([
        DailyTask(freelancer_id=a.id, task_date=date(2026, 8, 3), project_code="P-100", project_name="Project Alpha", task_description="Model", minutes_spent=120),
        DailyTask(freelancer_id=b.id, task_date=date(2026, 8, 4), project_code="P-100", project_name="Project Alpha", task_description="Coordinate", minutes_spent=90),
        DailyTask(freelancer_id=a.id, task_date=date(2026, 7, 30), project_code="P-100", project_name="Project Alpha", task_description="Earlier", minutes_spent=60),
        DailyTask(freelancer_id=b.id, task_date=date(2025, 6, 15), project_code="P-100", project_name="Project Alpha", task_description="Historical", minutes_spent=30),
    ])
    db.commit()
    return a, b, project


def test_release_version_and_no_new_schema_revision():
    config = (ROOT / "app" / "config.py").read_text(encoding="utf-8")
    assert 'APP_VERSION = "v3.0.24.3.1-release21.24.3.1-dashboard-leave-availability-hotfix"' in config
    versions = sorted((ROOT / "alembic" / "versions").glob("*.py"))
    assert any(path.name == "20260819_0018_freelancer_bank_details.py" for path in versions)
    assert not any("0019" in path.name for path in versions)


def test_project_time_by_member_follows_month_12m_and_all_time_periods():
    engine, db = _db()
    try:
        _seed_project_time(db)

        monthly = build_project_reports(db, period="month", month_key="2026-08")
        twelve = build_project_reports(db, period="12m", month_key="2026-08")
        all_time = build_project_reports(db, period="all", month_key="2026-08")

        assert monthly["project_time_by_member_rows"][0]["total_minutes"] == 210
        assert twelve["project_time_by_member_rows"][0]["total_minutes"] == 270
        assert all_time["project_time_by_member_rows"][0]["total_minutes"] == 300

        monthly_members = {m["name"]: m["minutes"] for m in monthly["project_time_by_member_rows"][0]["members"]}
        twelve_members = {m["name"]: m["minutes"] for m in twelve["project_time_by_member_rows"][0]["members"]}
        all_members = {m["name"]: m["minutes"] for m in all_time["project_time_by_member_rows"][0]["members"]}
        assert monthly_members == {"Member One": 120, "Member Two": 90}
        assert twelve_members == {"Member One": 180, "Member Two": 90}
        assert all_members == {"Member One": 180, "Member Two": 120}
    finally:
        db.close()
        engine.dispose()


def test_project_reports_page_export_carries_selected_period():
    template = (ROOT / "templates" / "project_reports.html").read_text(encoding="utf-8")
    assert "/portal/exports/project-work-time.xlsx?period={{ report.period }}&month={{ report.selected_month }}" in template
    assert "report.project_time_by_member_rows" in template
    assert "12-MONTH PROJECT WORK TIME" in template
    assert "ALL-TIME PROJECT WORK TIME" in template


def test_dedicated_project_work_time_export_matches_period_and_has_health_sheets():
    engine, db = _db()
    try:
        _seed_project_time(db)
        content = build_project_work_time_workbook(db, period="12m", month_key="2026-08", locale="en")
        wb = load_workbook(BytesIO(content), data_only=True)
        assert wb.sheetnames == ["Project Work Time Health", "Project Time by Member", "Monthly Breakdown"]
        member_sheet = wb["Project Time by Member"]
        assert member_sheet["A1"].value.startswith("PROJECT TIME BY MEMBER")
        rows = list(member_sheet.iter_rows(min_row=5, values_only=True))
        alpha = [row for row in rows if row[0] == "Project Alpha"]
        assert len(alpha) == 2
        assert {row[4]: row[6] for row in alpha} == {"Member One": "3h 00m", "Member Two": "1h 30m"}
        assert {row[3] for row in alpha} == {"4h 30m"}
    finally:
        db.close()
        engine.dispose()


def test_excel_exports_follow_portal_locale_for_english_and_traditional_chinese():
    engine, db = _db()
    try:
        _seed_project_time(db)
        en_bytes = build_project_work_time_workbook(db, period="month", month_key="2026-08", locale="en")
        zh_bytes = build_project_work_time_workbook(db, period="month", month_key="2026-08", locale="zh_TW")
        en = load_workbook(BytesIO(en_bytes), data_only=True)
        zh = load_workbook(BytesIO(zh_bytes), data_only=True)

        assert "Project Work Time Health" in en.sheetnames
        assert "專案工時概況" in zh.sheetnames
        assert en["Project Work Time Health"]["A4"].value == "Project"
        assert zh["專案工時概況"]["A4"].value == "專案"
        assert "2026年8月" in zh["專案工時概況"]["A1"].value

        # Existing export packages also localize their sheet names and headers
        # according to the same portal locale passed by the route.
        general_zh = load_workbook(BytesIO(build_export_workbook(db, month_key="2026-08", locale="zh_TW", include_tasks=True)), data_only=True)
        assert general_zh.sheetnames[0] == "匯出摘要"
        assert "所有任務" in general_zh.sheetnames
        assert general_zh["所有任務"]["A4"].value == "任務 ID"
    finally:
        db.close()
        engine.dispose()


def test_export_routes_pass_request_locale_and_support_period_parameter():
    source = (ROOT / "app" / "routers" / "portal.py").read_text(encoding="utf-8")
    assert 'from app.i18n import locale_for_request' in source
    assert '@router.get("/portal/exports/project-work-time.xlsx")' in source
    assert 'period: str = "month"' in source
    assert 'locale=locale_for_request(request)' in source
    assert 'EXPORT_PROJECT_WORK_TIME_XLSX' in source
